"""Đọc workbook sản lượng container của VPA.

Mỗi sheet là một tháng. Cột A là tên khu vực/cảng, cột E (`CỘNG`) là TEU
thông qua trong đúng tháng đó. Cột C là luỹ kế cùng kỳ năm trước, dùng để
suy ra từng tháng của năm không có sheet riêng (xem `cumulative_rows`).

Tên cảng trong nguồn không nhất quán giữa các kỳ - có kỳ mất dấu
(`Hai Phong`, `Đình vũ`), có kỳ thêm hậu tố (`(*)`, `(VPA)`, `/ cập nhật`,
`-Điều chỉnh số liệu T6/2024`). `normalize_name` gom hết về một khoá bỏ dấu
để mapping chỉ cần một dòng cho mỗi cảng.
"""
from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass

import openpyxl

# Sheet tháng có dạng T1-2026 / T01-2025. Sheet khác (Sheet1...) bỏ qua.
SHEET_RE = re.compile(r"^T(\d{1,2})-(\d{4})$")
# Header cột luỹ kế: "1T 2023", "12T2023", "6T-2025"...
CUM_RE = re.compile(r"^(\d{1,2})\s*T\s*[-\s]?\s*(\d{4})$")

NAME_COL, CUM_PREV_COL, MONTH_COL = 0, 2, 4
FIRST_DATA_ROW = 6
STOP_PREFIX = "TONG CONG"

# Hậu tố VPA gắn thêm vào tên cảng; không mang thông tin định danh.
_SUFFIXES = re.compile(
    r"\s*(\(\*\)|\(VPA\)|/.*|-\s*Điều chỉnh.*|\(TCOT cũ\))\s*$", re.IGNORECASE)


def strip_accents(text):
    # Đ/đ (U+0110/U+0111) is a distinct letter, not a base + combining mark, so
    # NFD leaves it alone and it has to be folded by hand - otherwise
    # "Đình Vũ" and "Dình Vũ" (a real typo in the source) stay different keys.
    folded = text.replace("Đ", "D").replace("đ", "d")
    return "".join(c for c in unicodedata.normalize("NFD", folded)
                   if unicodedata.category(c) != "Mn")


def normalize_name(raw):
    """Khoá tra cứu: bỏ hậu tố, bỏ dấu, gộp khoảng trắng, viết hoa."""
    text = " ".join(str(raw).split())
    while True:
        stripped = _SUFFIXES.sub("", text).strip()
        if stripped == text:
            break
        text = stripped
    return strip_accents(text).upper()


@dataclass(frozen=True)
class TeuRow:
    month: str          # "2025-06"
    name: str           # khoá đã chuẩn hoá
    raw_name: str       # nguyên văn trong file, để báo lỗi cho người đọc
    teu: float
    derived: bool       # True khi suy ra từ hiệu số luỹ kế


def _sheet_month(title):
    hit = SHEET_RE.match(title.strip())
    if not hit:
        return None
    month, year = int(hit.group(1)), int(hit.group(2))
    if not 1 <= month <= 12:
        raise ValueError(f"sheet {title!r} có tháng ngoài 1-12")
    return f"{year}-{month:02d}"


def _iter_name_rows(ws):
    """Sinh (raw_name, row) cho vùng dữ liệu, dừng ở dòng TỔNG CỘNG.

    Dưới TỔNG CỘNG là chú thích văn bản; đọc tiếp sẽ nuốt cả đoạn giải thích
    vào như thể là tên cảng.
    """
    for row in ws.iter_rows(min_row=FIRST_DATA_ROW, values_only=True):
        raw = row[NAME_COL]
        if raw is None or not str(raw).strip():
            continue
        raw = " ".join(str(raw).split())
        if normalize_name(raw).startswith(STOP_PREFIX):
            return
        yield raw, row


def _number(value):
    if value is None or isinstance(value, str):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def monthly_rows(path):
    """TEU từng tháng lấy trực tiếp từ các sheet tháng."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out = []
    try:
        for title in wb.sheetnames:
            month = _sheet_month(title)
            if month is None:
                continue
            for raw, row in _iter_name_rows(wb[title]):
                teu = _number(row[MONTH_COL] if len(row) > MONTH_COL else None)
                if teu is None:
                    continue
                out.append(TeuRow(month, normalize_name(raw), raw, teu, False))
    finally:
        wb.close()
    return out


def cumulative_rows(path):
    """TEU từng tháng của năm trước, suy ra bằng hiệu số hai cột luỹ kế.

    Sheet T01-2024 mang cột "1T 2023", T02-2024 mang "2T 2023"... nên tháng n
    của 2023 = luỹ kế n - luỹ kế (n-1). Số này là tự tính, không phải số VPA
    công bố, nên mọi bản ghi trả về đều mang derived=True.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    # {year: {n: {name: (cum, raw)}}}
    by_year = {}
    try:
        for title in wb.sheetnames:
            if _sheet_month(title) is None:
                continue
            ws = wb[title]
            header = next(ws.iter_rows(min_row=3, max_row=3, values_only=True))
            label = header[CUM_PREV_COL] if len(header) > CUM_PREV_COL else None
            hit = CUM_RE.match(" ".join(str(label or "").split()))
            if not hit:
                continue
            n, year = int(hit.group(1)), int(hit.group(2))
            slot = by_year.setdefault(year, {}).setdefault(n, {})
            for raw, row in _iter_name_rows(ws):
                cum = _number(row[CUM_PREV_COL] if len(row) > CUM_PREV_COL else None)
                if cum is not None:
                    slot[normalize_name(raw)] = (cum, raw)
    finally:
        wb.close()

    out = []
    for year, months in by_year.items():
        for n in sorted(months):
            prev = months.get(n - 1, {})
            for name, (cum, raw) in months[n].items():
                base = prev.get(name)
                # Tháng 1 lấy thẳng luỹ kế. Từ tháng 2, thiếu luỹ kế tháng
                # trước thì không suy được - bỏ, không đoán bằng 0 (sẽ thành
                # một tháng bằng cả luỹ kế nhiều tháng).
                if n == 1:
                    teu = cum
                elif base is None:
                    continue
                else:
                    teu = cum - base[0]
                out.append(TeuRow(f"{year}-{n:02d}", name, raw, teu, True))
    return out


def read_workbooks(paths, derive_years=()):
    """Gộp nhiều workbook. Số công bố luôn thắng số suy ra ở tháng trùng."""
    published, derived = [], []
    for path in paths:
        published.extend(monthly_rows(path))
        if derive_years:
            derived.extend(r for r in cumulative_rows(path)
                           if int(r.month[:4]) in derive_years)

    seen = {(r.month, r.name) for r in published}
    merged = list(published)
    merged.extend(r for r in derived if (r.month, r.name) not in seen)
    merged.sort(key=lambda r: (r.month, r.name))
    return merged
