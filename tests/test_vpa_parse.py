import openpyxl
import pytest

from scraper.vpa.parse import (cumulative_rows, monthly_rows, normalize_name,
                               read_workbooks)


def _sheet(wb, title, cum_label, rows):
    ws = wb.create_sheet(title)
    ws["A3"], ws["C3"] = "KHU VỰC / Cảng", cum_label
    ws["E4"] = "CỘNG"
    for i, (name, cum, teu) in enumerate(rows):
        ws.cell(row=6 + i, column=1, value=name)
        ws.cell(row=6 + i, column=3, value=cum)
        ws.cell(row=6 + i, column=5, value=teu)
    ws.cell(row=6 + len(rows), column=1, value="TỔNG CỘNG")
    return ws


@pytest.fixture()
def book(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _sheet(wb, "T01-2024", "1T 2023", [("Nam Đình Vũ", 100, 300),
                                       ("Đình vũ", 40, 90)])
    _sheet(wb, "T02-2024", "2T 2023", [("Nam Đình Vũ", 250, 320),
                                       ("Đình Vũ", 95, 80)])
    ws = wb.create_sheet("Sheet1")
    ws["A6"] = "ghi chú linh tinh"
    path = tmp_path / "vpa.xlsx"
    wb.save(path)
    return path


@pytest.mark.parametrize("raw,expected", [
    ("Đình vũ", "DINH VU"),
    ("Đình Vũ", "DINH VU"),
    ("Nam Hải Dình Vũ (*)", "NAM HAI DINH VU"),
    ("Nam Hải Đình Vũ", "NAM HAI DINH VU"),
    ("Hai Phong (Chùa Vẽ+Tân Vũ)", "HAI PHONG (CHUA VE+TAN VU)"),
    ("MIỀN BẮC (VPA)", "MIEN BAC"),
    ("VIMC Đình Vũ / cập nhật", "VIMC DINH VU"),
    ("Đồng Nai -Điều chỉnh số liệu T6/2024", "DONG NAI"),
    ("Tân Cảng - TCTT (TCOT cũ)", "TAN CANG - TCTT"),
    ("  Green  Port (*) ", "GREEN PORT"),
])
def test_normalize_name_folds_source_variants(raw, expected):
    assert normalize_name(raw) == expected


def test_monthly_rows_reads_the_month_column(book):
    rows = monthly_rows(book)

    jan = {r.name: r.teu for r in rows if r.month == "2024-01"}
    assert jan == {"NAM DINH VU": 300, "DINH VU": 90}
    assert all(r.derived is False for r in rows)


def test_monthly_rows_ignores_non_month_sheets(book):
    assert not [r for r in monthly_rows(book) if "ghi chú" in r.raw_name]


def test_parsing_stops_at_the_total_row(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _sheet(wb, "T01-2024", "1T 2023", [("Nam Đình Vũ", 1, 10)])
    ws = wb["T01-2024"]
    ws.cell(row=8, column=1, value="Kể từ năm 2012, sản lượng thông qua...")
    ws.cell(row=8, column=5, value=999)
    path = tmp_path / "v.xlsx"
    wb.save(path)

    assert [r.name for r in monthly_rows(path)] == ["NAM DINH VU"]


def test_cumulative_rows_difference_consecutive_months(book):
    rows = {(r.month, r.name): r.teu for r in cumulative_rows(book)}

    assert rows[("2023-01", "NAM DINH VU")] == 100      # tháng 1 = luỹ kế
    assert rows[("2023-02", "NAM DINH VU")] == 150      # 250 - 100
    assert rows[("2023-02", "DINH VU")] == 55           # 95 - 40, bất kể dấu


def test_cumulative_rows_are_flagged_as_derived(book):
    assert all(r.derived for r in cumulative_rows(book))


def test_read_workbooks_prefers_published_over_derived(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    # Sheet tháng của 2023 và cột luỹ kế 2023 cùng tồn tại, giá trị khác nhau.
    _sheet(wb, "T01-2023", "1T 2022", [("Nam Đình Vũ", 5, 111)])
    _sheet(wb, "T01-2024", "1T 2023", [("Nam Đình Vũ", 999, 300)])
    path = tmp_path / "v.xlsx"
    wb.save(path)

    rows = [r for r in read_workbooks([path], derive_years=(2023,))
            if r.month == "2023-01"]

    assert [(r.teu, r.derived) for r in rows] == [(111, False)]


def test_read_workbooks_skips_derivation_when_not_asked(book):
    assert not [r for r in read_workbooks([book]) if r.derived]
